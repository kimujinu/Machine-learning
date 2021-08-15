# pip3 install openpyxl
import openpyxl

book = openpyxl.load_workbook("test_sample.xlsx")

# 첫 번째 방법
#print(book.get_sheet_names())
#print(book.get_sheet_by_name("Sheet1"))

# 두 번째 방법
sheet = book.worksheets[0]
for row in sheet.rows:
    print()
    for data in row:
        print(data.value,end=' ')


workbook = openpyxl.Workbook()
sheet = workbook.active

sheet["A1"] = "업무자동화"
sheet["B2"] = "시작"
sheet.merge_cells("A1:C1")
sheet["A1"].font = openpyxl.styles.Font(size=20,color="FF0000")

workbook.save("newFile.xlsx")